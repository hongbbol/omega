#!/usr/bin/env python3
"""
고양이 오메가3 제품 채점 엔진 (portable)
- 입력: products.csv  (제품 데이터 — 이 파일만 편집하면 됨)
- 출력: 고양이_오메가3_제품_채점표.xlsx  (기준 시트 + 점수표 시트, 수식 자동 채점/순위)

사용법:
    pip install openpyxl
    python build_xlsx.py
    # 생성된 xlsx를 Excel 또는 LibreOffice로 열면 수식이 자동 계산됩니다.

새 제품 추가/수정:
    products.csv 에 행을 추가하거나 값을 채우세요.
    카테고리 컬럼은 아래 RUBRIC의 허용 토큰만 사용해야 점수가 매겨집니다.
    숫자 컬럼(밀도/순도)은 숫자만, 모르면 빈칸으로 두세요(=0점 처리, 가정 입력 금지).
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CSV_IN  = "products.csv"
XLSX_OUT = "고양이_오메가3_제품_채점표.xlsx"

# ===== RUBRIC (채점 기준) =====================================================
# 상위 가중치: 성분 35 / 산패 30 / 원료·형태 20 / 투명성 15  (합 100)
# 안전 점수상한: 비해양성(ALA단독)→20, 대구간유→40, 산패고위험→50
#   ※조류(algae) 오일은 EPA/DHA 직접공급 = 정상 / 폴락 어체유는 간유 아님
# 빈칸/미상 처리: 공개 안 한 항목은 해당 세부점수 0 (제외 아님). 가정 입력 금지.
LOOKUPS = {
    "비율표기": [("명시",5),("일부",2),("미표기",0),("미상",0)],
    "형태":     [("rTG",10),("TG",10),("EE",5),("미상",5)],
    "어종":     [("소형어",10),("일반어",6),("식물성",0),("조류",8),("미상",3)],
    "항산화제": [("있음",10),("없음",0),("미상",0)],
    "포장":     [("우수",10),("보통",5),("취약",0),("미상",0)],
    "제형":     [("캡슐",10),("소프트젤",10),("츄",6),("스틱",6),("펌프액상",6),("대용량액상",2),("미상",0)],
    "인증":     [("있음",5),("없음",0),("미상",0)],
    "COA":      [("공개",6),("일부",3),("없음",0),("미상",0)],
    "원료사":   [("명시",4),("일부",2),("없음",0),("미상",0)],
}
# 밀도 구간(EPA+DHA mg/단위): <100→5 / 100-199→10 / 200-299→15 / ≥300→20 / 빈칸→0
# 순도 구간(%):                <40→3 / 40-59→6 / ≥60→10 / 빈칸→0
# =============================================================================

HDR=Font(bold=True,color="FFFFFF",name="Arial"); HFILL=PatternFill("solid",fgColor="06B470")
BOLD=Font(bold=True,name="Arial"); NORM=Font(name="Arial")

def num(v):
    try:
        return float(v) if str(v).strip()!="" else None
    except ValueError:
        return None

def build():
    rows=list(csv.DictReader(open(CSV_IN,encoding="utf-8-sig")))
    wb=Workbook()

    # ---- 기준 시트 (룩업 테이블) ----
    r=wb.active; r.title="기준"
    r["A1"]="고양이 오메가3 제품 채점 기준 (공개정보 기반·실측 전)"; r["A1"].font=Font(bold=True,size=13,name="Arial")
    r["A2"]=("가중치 성분35/산패30/원료·형태20/투명성15 | 안전상한 비해양성(ALA)→20·대구간유→40·산패고위험→50 "
             "| 조류오일=정상, 폴락=어체유(간유 아님)"); r["A2"].font=Font(italic=True,name="Arial")
    RNG={}; row0=4
    for name,table in LOOKUPS.items():
        r.cell(row0-1,2,name).font=BOLD
        for i,(k,v) in enumerate(table):
            r.cell(row0+i,2,k).font=NORM; r.cell(row0+i,3,v).font=NORM
        RNG[name]=f"기준!$B${row0}:$C${row0+len(table)-1}"
        row0+=len(table)+2
    r.column_dimensions["B"].width=16; r.column_dimensions["C"].width=8

    # ---- 점수표 시트 ----
    s=wb.create_sheet("점수표")
    head=["번호","제품명","분류","밀도(EPA+DHA mg/단위)","순도(%)","비율표기","형태","어종","항산화제","포장","제형",
          "인증","COA","원료사","비해양성","대구","산패고위험",
          "성분/35","산패/30","원료형태/20","투명성/15","합계","안전상한","최종점수","잠정순위",
          "정보검증","적합성","데이터충실도","출처·비고"]
    for j,h in enumerate(head,1):
        c=s.cell(1,j,h); c.font=HDR; c.fill=HFILL
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    thin=Side(style="thin",color="DDDDDD"); bd=Border(thin,thin,thin,thin)
    f0=2; n=len(rows)
    for i,row in enumerate(rows):
        rr=f0+i
        vals=[row["번호"],row["제품명"],row["분류"],num(row["밀도_EPADHA_mg"]),num(row["순도_pct"]),
              row["비율표기"],row["형태"],row["어종"],row["항산화제"],row["포장"],row["제형"],
              row["인증"],row["COA"],row["원료사"],row["비해양성"],row["대구"],row["산패고위험"]]
        for j,v in enumerate(vals,1):
            c=s.cell(rr,j,v); c.font=NORM; c.border=bd
            if j>=4: c.alignment=Alignment(horizontal="center")
        # 점수 수식
        comp = (f'IF(NOT(ISNUMBER(D{rr})),0,IF(D{rr}>=300,20,IF(D{rr}>=200,15,IF(D{rr}>=100,10,5))))'
                f'+IF(NOT(ISNUMBER(E{rr})),0,IF(E{rr}>=60,10,IF(E{rr}>=40,6,3)))'
                f'+IFERROR(VLOOKUP(F{rr},{RNG["비율표기"]},2,FALSE),0)')
        sanpae=(f'IFERROR(VLOOKUP(I{rr},{RNG["항산화제"]},2,FALSE),0)+IFERROR(VLOOKUP(J{rr},{RNG["포장"]},2,FALSE),0)'
                f'+IFERROR(VLOOKUP(K{rr},{RNG["제형"]},2,FALSE),0)')
        wonryo=(f'IFERROR(VLOOKUP(G{rr},{RNG["형태"]},2,FALSE),0)+IFERROR(VLOOKUP(H{rr},{RNG["어종"]},2,FALSE),0)')
        trans =(f'IFERROR(VLOOKUP(L{rr},{RNG["인증"]},2,FALSE),0)+IFERROR(VLOOKUP(M{rr},{RNG["COA"]},2,FALSE),0)'
                f'+IFERROR(VLOOKUP(N{rr},{RNG["원료사"]},2,FALSE),0)')
        s.cell(rr,18,"="+comp); s.cell(rr,19,"="+sanpae); s.cell(rr,20,"="+wonryo); s.cell(rr,21,"="+trans)
        s.cell(rr,22,f"=SUM(R{rr}:U{rr})")
        s.cell(rr,23,f'=MIN(IF(O{rr}="Y",20,100),IF(P{rr}="Y",40,100),IF(Q{rr}="Y",50,100))')
        s.cell(rr,24,f"=MIN(V{rr},W{rr})")
        s.cell(rr,25,f"=RANK(X{rr},$X${f0}:$X${f0+n-1},0)")
        s.cell(rr,26,row["정보검증"]); s.cell(rr,27,row["적합성"])
        comp_cnt=("+".join([f'IF(ISNUMBER(D{rr}),1,0)',f'IF(ISNUMBER(E{rr}),1,0)']
                  +[f'IF({col}{rr}<>"미상",1,0)' for col in "FGHIJKLMN"]))
        s.cell(rr,28,f'=({comp_cnt})&"/11"')
        s.cell(rr,29,row["출처비고"])
        for j in range(18,30):
            c=s.cell(rr,j); c.font=NORM; c.border=bd
            if j<28: c.alignment=Alignment(horizontal="center")
    widths={"A":5,"B":26,"C":13,"D":13,"E":8,"F":9,"G":7,"H":8,"I":9,"J":8,"K":11,"L":7,"M":7,"N":8,
            "O":8,"P":6,"Q":10,"R":8,"S":8,"T":11,"U":10,"V":7,"W":9,"X":9,"Y":9,"Z":11,"AA":9,"AB":11,"AC":44}
    for k,v in widths.items(): s.column_dimensions[k].width=v
    s.freeze_panes="D2"; s.row_dimensions[1].height=42
    fill=PatternFill("solid",fgColor="EAF7F0")
    for rr in range(f0,f0+n):
        for j in range(18,26): s.cell(rr,j).fill=fill
    wb.save(XLSX_OUT)
    print(f"생성 완료: {XLSX_OUT}  (제품 {n}개)")
    print("→ Excel/LibreOffice로 열면 수식이 자동 계산됩니다.")

if __name__=="__main__":
    if not os.path.exists(CSV_IN):
        raise SystemExit(f"{CSV_IN} 가 없습니다. 같은 폴더에 두고 실행하세요.")
    build()
